"""Run calibration centiloid pipeline for PiB"""
__author__ = "Pawel J Markiewicz"
__copyright__ = "Copyright 2022"
import pickle
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import openpyxl as xl
from scipy.stats import linregress

import amypet
from amypet.centiloid import run as centiloid_run
from amypet.utils import cl_anchor_fldr

# > input paths
drv = Path('/data/AMYPET')
atlases = drv / 'Atlas' / 'CL_2mm'
opth = drv / 'CL' / 'PiB'

# ----------------------------------------------------------------------
# AD
dirpet = drv / 'CL' / 'PiB' / 'AD-100_PET_5070' / 'nifti'
dirmri = drv / 'CL' / 'PiB' / 'AD-100_MR' / 'nifti'
fpets = sorted(dirpet.glob("*.*"))
fmris = sorted(dirmri.glob("*.*"))

# > run visual check of the images before running the CL pipeline
# amypet.im_check_pairs(fpets, fmris)

flip_pet = len(fpets) * [(1, 1, 1)]
flip_pet[9] = (1, -1, 1)
flip_pet[22] = (1, -1, 1)
flip_pet[26] = (1, -1, 1)
flip_pet[34] = (1, -1, 1)
flip_pet[36] = (1, -1, 1)
flip_pet[40] = (1, -1, 1)
flip_pet[44] = (1, -1, 1)

out_ad = centiloid_run(fpets, fmris, atlases, flip_pet=flip_pet, outpath=opth / 'output_pib_ad')
with open(str(opth / 'output_pib_ad.pkl'), 'wb') as f:
    pickle.dump(out_ad, f)
# ----------------------------------------------------------------------

# ----------------------------------------------------------------------
# YC
dirpet = drv / 'CL' / 'PiB' / 'YC-0_PET_5070' / 'nifti'
dirmri = drv / 'CL' / 'PiB' / 'YC-0_MR' / 'nifti'
fpets = sorted(dirpet.glob("*.*"))
fmris = sorted(dirmri.glob("*.*"))

# > run visual check of the images before running the CL pipeline
# amypet.im_check_pairs(fpets, fmris)

out_yc = centiloid_run(fpets, fmris, atlases, outpath=opth / 'output_pib_yc')
with open(str(opth / 'output_pib_yc.pkl'), 'wb') as f:
    pickle.dump(out_yc, f)
# ----------------------------------------------------------------------

# ----------------------------------------------------------------------
with open(str(opth / 'output_pib_yc.pkl'), 'rb') as f:
    out_yc = pickle.load(f)

with open(str(opth / 'output_pib_ad.pkl'), 'rb') as f:
    out_ad = pickle.load(f)
# ----------------------------------------------------------------------

refs = amypet.get_clref(opth / 'SupplementaryTable1.xlsx')

diff = amypet.check_suvrs(out_yc, out_ad, refs)

diff = amypet.check_cls(out_yc, out_ad, diff, refs)

# cla = amypet.save_cl_anchors(diff, outpath=Path('/home/pawel/Dropbox'))
cla = amypet.save_cl_anchors(diff)

# > load/test the anchor and conversion tables
cl_fldr = cl_anchor_fldr
cla = amypet.get_cl_anchors(cl_fldr)
cla = amypet.get_cl_anchors(cl_fldr.parent)

tracer = 'flute'
tracer = 'fbb'
tracer = 'fbp'
clc = amypet.get_suvr2pib(tracer, path=cl_fldr)
clc = amypet.get_suvr2pib(tracer, path=cl_fldr.parent)


# ----------------------------------------------------------------------
def identity_line(ax=None, ls='--', *args, **kwargs):
    '''
    plot identity line for any correlation analysis
    needed for the CL calibration and beyond.
    '''
    # see: https://stackoverflow.com/q/22104256/3986320
    ax = ax or plt.gca()
    identity, = ax.plot([], [], ls=ls, *args, **kwargs)

    def callback(axes):
        low_x, high_x = ax.get_xlim()
        low_y, high_y = ax.get_ylim()
        low = min(low_x, low_y)
        high = max(high_x, high_y)
        identity.set_data([low, high], [low, high])

    callback(ax)
    ax.callbacks.connect('xlim_changed', callback)
    ax.callbacks.connect('ylim_changed', callback)
    return ax


# ===================================================
drv = Path('/Users/pawel/Dropbox/tmp/AMYPAD/AmyPET/')

# - - - - - - - - - - - - - - - - - - - - - - - - -
with open(drv / 'cal_fbb.pkl', 'rb') as f:
    cal = pickle.load(f)

info = xl.load_workbook(drv / 'FBBproject_SupplementaryTable.xlsx')
dat = info['18F-FBB']
cl_o = np.array([i.value for i in dat['H'][4:29]] + [i.value for i in dat['H'][30:]])

pid = [i.value for i in dat['A'][4:29]] + [int(i.value[1:]) for i in dat['A'][30:]]
idxs = [pid.index(int(i)) for i in cal['wc']['sbj']]
cl_os = cl_o[idxs]

rvoi = 'wc'
cl_amy = cal[rvoi]['calib']['cl_std_fbb']

fig, ax = plt.subplots()
ax.scatter(cl_os, cl_amy, c='black')
identity_line(ax=ax, ls='--', c='b')
ax.set_xlabel('Original F18 CLs')
ax.set_ylabel('AmyPET F18 CLs')
ax.grid('on')
ax.set_title('FBB')

m, a, r, p, stderr = linregress(cl_os, cl_amy)
r2 = r**2
ax.text(0, 125, f'$R^2={r2:.4f}$', fontsize=12)
# - - - - - - - - - - - - - - - - - - - - - - - - -

# - - - - - - - - - - - - - - - - - - - - - - - - -
with open(drv / 'cal_flt.pkl', 'rb') as f:
    cal = pickle.load(f)

info = xl.load_workbook(drv / 'flutemetamol.xlsx')
dat = info['GE PiB & Flutemetamol']
cl_o = np.array([i.value for i in dat['T'][3:53]] + [i.value for i in dat['T'][54:78]])

pid = [int(i.value[3:]) for i in dat['C'][3:53]] + [int(i.value[4:]) for i in dat['C'][54:78]]

idxs = [pid.index(int(i)) for i in cal['wc']['sbj']]
cl_os = cl_o[idxs]

rvoi = 'wc'
cl_amy = cal[rvoi]['calib']['cl_std_fbb']

fig, ax = plt.subplots()
ax.scatter(cl_os, cl_amy, c='black')
identity_line(ax=ax, ls='--', c='b')
ax.set_xlabel('Original F18 CLs')
ax.set_ylabel('AmyPET F18 CLs')
ax.grid('on')
ax.set_title('FLUTE')

m, a, r, p, stderr = linregress(cl_os, cl_amy)
r2 = r**2
ax.text(0, 125, f'$R^2={r2:.4f}$', fontsize=12)
# - - - - - - - - - - - - - - - - - - - - - - - - -

# - - - - - - - - - - - - - - - - - - - - - - - - -
with open(drv / 'cal_fbp.pkl', 'rb') as f:
    cal = pickle.load(f)

info = xl.load_workbook(drv / 'Avid_Centiloid_standard_method.xlsx')
dat = info['Sheet1']
cl_o = np.array([i.value for i in dat['I'][2:]])

pid = [i.value for i in dat['A'][2:]]

idxs = [pid.index(int(i)) for i in cal['wc']['sbj']]
cl_os = cl_o[idxs]

rvoi = 'wc'
cl_amy = cal[rvoi]['calib']['cl_std_fbb']

fig, ax = plt.subplots()
ax.scatter(cl_os, cl_amy, c='black')
identity_line(ax=ax, ls='--', c='b')
ax.set_xlabel('Original F18 CLs')
ax.set_ylabel('AmyPET F18 CLs')
ax.grid('on')
ax.set_title('FBP')

m, a, r, p, stderr = linregress(cl_os, cl_amy)
r2 = r**2
ax.text(0, 125, f'$R^2={r2:.4f}$', fontsize=12)
