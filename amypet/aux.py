'''
Auxiliary functions for the centiloid project
'''
import os
import pickle
import re
import numpy as np
import openpyxl
import matplotlib.pyplot as plt
from scipy.stats import linregress

from niftypet import nimpa

# > regions used in CL project
rvois = ['wc', 'cg', 'wcb', 'pns']

# > region full name strings for plots
rvoi_str = dict(
    wc='whole cerebellum',
    cg='cerebellum GM',
    wcb='whole cerebellum + brain stem',
    pns='pons')

#----------------------------------------------------------------------
def im_check_pairs(fpets, fmris):
    '''
    checks visually image by image if the PET and MR have
    the same orientation.
    '''
    fig,ax = plt.subplots(2,2,figsize=(10,10))
    for fp, fm in zip(fpets, fmris):
        print(f'{fp} : {fm}')
        p = nimpa.getnii(fp)
        m = nimpa.getnii(fm)

        ps = p.shape
        ms = m.shape

        ax[0,0].imshow(p[ps[0]//2, ...], cmap='magma')
        ax[1,0].imshow(m[ms[0]//2, ...], cmap='bone')

        ax[0,1].imshow(p[..., ps[2]//2], cmap='magma')
        ax[1,1].imshow(m[..., ms[2]//2], cmap='bone')

        plt.draw()
        plt.waitforbuttonpress(0)
#----------------------------------------------------------------------


#----------------------------------------------------------------------
def identity_line(ax=None, ls='--', *args, **kwargs):
    '''
    plot identity line for any correlation analysis
    needed for the CL calibration and beyond.
    '''
    # see: https://stackoverflow.com/q/22104256/3986320
    ax = ax or plt.gca()
    identity, = ax.plot([], [], ls=ls, *args, **kwargs)
    def callback(axes):
        low_x, high_x = ax.get_xlim()
        low_y, high_y = ax.get_ylim()
        low = min(low_x, low_y)
        high = max(high_x, high_y)
        identity.set_data([low, high], [low, high])
    callback(ax)
    ax.callbacks.connect('xlim_changed', callback)
    ax.callbacks.connect('ylim_changed', callback)
    return ax
#----------------------------------------------------------------------


#----------------------------------------------------------------------
def get_clref(fxls):
    '''
    get the excel sheets with CL reference data from the CL project
    excel sheet.
    '''
    
    # > Define variable to load the workbook
    wb = openpyxl.load_workbook(fxls)
    ws = wb.active

    pib_tbl = dict(yc={},ad={})

    # > cell offsets
    ioff_ad = [5,50]
    ioff_yc = [51,85]

    sbj = [str(i.value) for i in ws['A']]
    cg  = [str(i.value) for i in ws['B']]
    wc  = [str(i.value) for i in ws['C']]
    wcb = [str(i.value) for i in ws['D']]
    pns = [str(i.value) for i in ws['E']]

    ccg  = [str(i.value) for i in ws['F']]
    cwc  = [str(i.value) for i in ws['G']]
    cwcb = [str(i.value) for i in ws['H']]
    cpns = [str(i.value) for i in ws['I']]


    # > young control indices and SUVr/CL values
    pib_tbl['yc']['id'] = np.array([int(i[3:]) for i in sbj[ioff_yc[0]:ioff_yc[1]]])

    pib_tbl['yc']['suvr'] = {}
    pib_tbl['yc']['suvr']['cg']  = np.array([float(i) for i in  cg[ioff_yc[0]:ioff_yc[1]]])
    pib_tbl['yc']['suvr']['wc']  = np.array([float(i) for i in  wc[ioff_yc[0]:ioff_yc[1]]])
    pib_tbl['yc']['suvr']['wcb'] = np.array([float(i) for i in wcb[ioff_yc[0]:ioff_yc[1]]])
    pib_tbl['yc']['suvr']['pns'] = np.array([float(i) for i in pns[ioff_yc[0]:ioff_yc[1]]])

    pib_tbl['yc']['cl'] = {}
    pib_tbl['yc']['cl']['cg']  = np.array([float(i) for i in  ccg[ioff_yc[0]:ioff_yc[1]]])
    pib_tbl['yc']['cl']['wc']  = np.array([float(i) for i in  cwc[ioff_yc[0]:ioff_yc[1]]])
    pib_tbl['yc']['cl']['wcb'] = np.array([float(i) for i in cwcb[ioff_yc[0]:ioff_yc[1]]])
    pib_tbl['yc']['cl']['pns'] = np.array([float(i) for i in cpns[ioff_yc[0]:ioff_yc[1]]])

    # > AD indices and SUVr/CL
    pib_tbl['ad']['id'] = np.array([int(i[3:]) for i in sbj[ioff_ad[0]:ioff_ad[1]]])
    pib_tbl['ad']['suvr'] = {}
    pib_tbl['ad']['suvr']['cg']  = np.array([float(i) for i in  cg[ioff_ad[0]:ioff_ad[1]]])
    pib_tbl['ad']['suvr']['wc']  = np.array([float(i) for i in  wc[ioff_ad[0]:ioff_ad[1]]])
    pib_tbl['ad']['suvr']['wcb'] = np.array([float(i) for i in wcb[ioff_ad[0]:ioff_ad[1]]])
    pib_tbl['ad']['suvr']['pns'] = np.array([float(i) for i in pns[ioff_ad[0]:ioff_ad[1]]])

    pib_tbl['ad']['cl'] = {}
    pib_tbl['ad']['cl']['cg']  = np.array([float(i) for i in  ccg[ioff_ad[0]:ioff_ad[1]]])
    pib_tbl['ad']['cl']['wc']  = np.array([float(i) for i in  cwc[ioff_ad[0]:ioff_ad[1]]])
    pib_tbl['ad']['cl']['wcb'] = np.array([float(i) for i in cwcb[ioff_ad[0]:ioff_ad[1]]])
    pib_tbl['ad']['cl']['pns'] = np.array([float(i) for i in cpns[ioff_ad[0]:ioff_ad[1]]])

    return pib_tbl
    #----------------------------------------------------------------------




def check_suvrs(suvr_yc, suvr_ad, refs):
    ''' Check/QC the obtained SUVRs in `suvr_yc` and `suvr_ad `dictionaries
        in relation to the reference in `refs` dictionary for the groups
        ('ad' or 'yc').

        Returns dictionary with the mean SUVrs and differences.
    '''

    # > prepare diff dictionary for the differences observed
    # > in young controls (yc) and AD patients (ad)
    diff = dict(yc={}, ad={})

    # > initialise the means for each reference VOI
    for rvoi in rvois:
        diff['yc'][rvoi] = dict(mean_ref=0, mean=0, N=0)
        diff['ad'][rvoi] = dict(mean_ref=0, mean=0, N=0)


    def run_checks(suvr_dct, grp):

        for rvoi in rvois:
            print('========================================================')
            for k in suvr_dct:

                if grp=='yc':
                    idx = int(k[2:5])
                elif grp=='ad':
                    idx = int(k[2:4])
                else:
                    raise ValueError('e> unknown group - only <yc>  or <ad> are accepted')
                
                i = np.where(refs[grp]['id']==idx)[0][0]

                suvr = suvr_dct[k]['suvr'][rvoi]
                suvr_ref = refs[grp]['suvr'][rvoi][i]
                err = 100*(suvr-suvr_ref)/suvr_ref

                diff[grp][rvoi]['N'] += 1
                diff[grp][rvoi]['mean_ref'] += suvr_ref
                diff[grp][rvoi]['mean'] += suvr
                diff[grp][rvoi][k] = dict(suvr=suvr, ref=suvr_ref, err=err)
                
                print(f'refvoi={rvoi}, indx> {idx}, suvr={suvr:.3f}, ref={suvr_ref:.3f}, error={err:.3f}%')

            diff[grp][rvoi]['mean'] /= diff[grp][rvoi]['N']
            diff[grp][rvoi]['mean_ref'] /= diff[grp][rvoi]['N']

            # relative % mean difference
            emean = diff[grp][rvoi]['mean']
            rmean = diff[grp][rvoi]['mean_ref']
            rmd = (emean-rmean)/rmean
            diff[grp][rvoi]['mean_diff'] = rmd
            print('--------------------------------------------------------')
            print(f'> group % mean difference: {rmd:.3f}% (suvr={emean:.3f}, ref={rmean:.3f})')

    print('========================================================')
    print('YOUNG CONTROLS (YC)')
    print('========================================================')
    run_checks(suvr_yc, 'yc')

    print('========================================================')
    print('AD PATIENTS (AD)')
    print('========================================================')
    run_checks(suvr_ad, 'ad')

    return diff
#----------------------------------------------------------------------




#----------------------------------------------------------------------
def check_cls(suvr_yc, suvr_ad, diff, refs):
    ''' check centiloid values relative to the reference in `refs` 
        dictionary, using SUVr dictionaries for YC and AD groups
        `suvr_yc` and `suvr_ad`.  Also, update the `diff` dictionary
        with the centiloid results and output it. 
    '''

    for rvoi in rvois:
        print('=====================================================')
        suvr_0 = diff['yc'][rvoi]['mean']
        suvr_100 = diff['ad'][rvoi]['mean']

        grp = 'yc'
        suvr_dct = suvr_yc

        for k in suvr_dct:

            idx = int(k[2:5])
            i = np.where(refs[grp]['id']==idx)[0][0]

            cl = 100*(suvr_dct[k]['suvr'][rvoi] - suvr_0)/(suvr_100-suvr_0)
            cl_ref = refs[grp]['cl'][rvoi][i]

            diff[grp][rvoi][k].update(dict(cl=cl, cl_ref=cl_ref))

            print(f'refvoi={rvoi}, indx> {idx}, cl={cl:.3f}, ref={cl_ref:.3f}')

        grp = 'ad'
        suvr_dct = suvr_ad

        for k in suvr_dct:

            idx = int(k[2:4])
            i = np.where(refs[grp]['id']==idx)[0][0]

            cl = 100*(suvr_dct[k]['suvr'][rvoi] - suvr_0)/(suvr_100-suvr_0)
            cl_ref = refs[grp]['cl'][rvoi][i]

            diff[grp][rvoi][k].update(dict(cl=cl, cl_ref=cl_ref))

            print(f'refvoi={rvoi}, indx> {idx}, cl={cl:.3f}, ref={cl_ref:.3f}')



    def get_cls(diff, grp, rvoi, cl='cl'):
        ''' Get centiloid values in a list for any group and reference VOI.
        '''
        return [diff[grp][rvoi][k][cl] for k in diff[grp][rvoi] if grp in k.lower()]

    #------------------------------------------------------------------
    fig, ax = plt.subplots(2,2, figsize=(10,10))

    for a, rvoi in enumerate(rvois):
        i = a//2
        j = a%2

        # > combine CLs for both groups
        cl = get_cls(diff, 'yc', rvoi, cl='cl')
        cl += get_cls(diff, 'ad', rvoi, cl='cl')

        # > combine reference CLs into one list
        clref = get_cls(diff, 'yc', rvoi, cl='cl_ref')
        clref += get_cls(diff, 'ad', rvoi, cl='cl_ref')

        # > find the linear regression between the CL and reference CL
        m, a, r, p, stderr = linregress(clref, cl)
        r2 = r**2
        
        # > plot the scatter plots for all 4 reference VOIs
        ax[i,j].scatter(clref, cl, c='black')
        identity_line(ax=ax[i,j], ls='--', c='b')
        ax[i,j].text(60, 20, f'$y = {m:.4f}x + {a:.4f}$', fontsize=12)
        ax[i,j].text(60, 10, f'$R^2={r2:.4f}$', fontsize=12)
        ax[i,j].set_title(rvoi_str[rvoi])
        ax[i,j].set_xlabel('CL ref')
        ax[i,j].set_ylabel('CL AmyPET')
        ax[i,j].grid('on')

    fig.tight_layout()
    #------------------------------------------------------------------

    return diff
#----------------------------------------------------------------------


def save_cl_anchors(diff):
    ''' save the CL anchor points for each reference VOI.
    '''

    CLA = {}
    txt = '# centiloid anchor points for different reference VOIs\n'
    for rvoi in rvois:
        suvr_0 = diff['yc'][rvoi]['mean']
        suvr_100 = diff['ad'][rvoi]['mean']
        CLA[rvoi] = (suvr_0, suvr_100)
        txt += f'cla_{rvoi} = ({suvr_0}, {suvr_100})\n'
    print(txt)
    
    cpth = os.path.realpath(__file__)
    pth = os.path.join(os.path.dirname(cpth), 'CL_PiB_anchors.pkl')

    with open(pth, 'wb') as f:
        pickle.dump(CLA, f)

    CLA['path'] = pth
    
    return CLA


def get_cl_anchors():
    ''' load the centiloid anchor points
    '''

    cpth = os.path.realpath(__file__)
    pth = os.path.join(os.path.dirname(cpth), 'CL_PiB_anchors.pkl')


    with open(pth, 'rb') as f:
        CLA = pickle.load(f)

    return CLA



#======================================================================
# CL CALIBRATION FOR A NEW TRACER
#======================================================================

def calib_tracer(
    outpib,
    outnew,
    xystr=None,
    figsize=(10,10),
    fontsize=12):
    '''calibrate CL scale for a new amyloid PET tracer,
       given the dictionaries from processing the PiB 
       and a new tracer scans, i.e., `outpib` and `outnew`.

       Other arguments:
       xystr - dictionary of plot string XY locations.
       figsize - tuple for controlling the figure size
    '''

    # > get PiB centiloid anchor points
    clap = get_cl_anchors()

    if xystr is None:
        xystr = dict(wc=[1.5, 1.1], cg=[2.0, 1.2], wcb=[1.5, 1.0], pns=[1.0, 0.75])

    # > extracting the index from names
    pp = re.compile('(GE_\w*_\w*_(\d*)_NIFTI)|(Y*(\d*)\w*)')

    # > calibration dictionary with SUVr and CL values
    cal = {}
    for rvoi in rvois:

        print('------------------------------------------------------------')
        print(rvoi_str[rvoi])

        # > PiB and F18 SUVrs and PiB CL in one list/array
        cl_suvr = []

        # > initialise dictionary
        cal[rvoi] = dict(sbj={}, calib={})

        for k in outpib:


            idx = pp.match(k)[2] or pp.match(k)[4]

            # > get the index in the NEW tracer dataset
            kf = [ky for ky in outnew if idx+'_' in ky]
            if len(kf)==1:
                kf = kf[0]
            else:
                raise ValueError('e> problem with F18 index...')
            
            # > get the SUVr for PiB and FBB
            suvrp = outpib[k]['suvr'][rvoi]
            suvrf = outnew[kf]['suvr'][rvoi]

            # > calculate the CL for the calibrating PiB scans
            cl = 100*(suvrp - clap[rvoi][0])/(clap[rvoi][1]-clap[rvoi][0])

            # > put the SUVrs for PiB and FBB into dictionary
            cal[rvoi]['sbj'][idx] = dict(
                cl=cl,
                suvrp=suvrp,
                suvrf=suvrf)

            # > ... and a list for a future array
            cl_suvr.append([cl, cal[rvoi]['sbj'][idx]['suvrp'], cal[rvoi]['sbj'][idx]['suvrf']])

            print(f'refvoi={rvoi}, indx> {idx}, cl={cl:.3f}, suvr_pib={suvrp:.3f}, suvr_fbb={suvrf:.3f}')

        
        #------------------------------------------------------------------
        # > find the linear relationship for FBB_SUVr = m*PiB_SUVr + b
        # > calculate the tracer `m_std` and `b_std` 
        # > (Eq. 2.2.3.1a in Klunk et al. 2015)

        cl_suvr = np.array(cl_suvr)
        m_std, b_std, r, p, stderr = linregress(cl_suvr[:,1], cl_suvr[:,2])
        r2 = r**2

        suvr_pib_calc =  (cl_suvr[:,2]-b_std) / m_std

        cl_std_fbb = 100*(suvr_pib_calc - clap[rvoi][0])/(clap[rvoi][1]-clap[rvoi][0])

        cal[rvoi]['calib'] = dict(
            m_std=m_std,
            b_std=b_std,
            r2=r2,
            p=p,
            stderr=stderr,
            cl_suvr=cl_suvr,
            suvr_pib_calc=suvr_pib_calc,
            cl_std_fbb=cl_std_fbb)
        #------------------------------------------------------------------

        
    #----------------------------------------------------------------------
    # VISUALISATION

    fig,  ax  = plt.subplots(2,2, figsize=figsize)
    fig2, ax2 = plt.subplots(2,2, figsize=figsize)

    for a, rvoi in enumerate(rvois):
        i = a//2
        j = a%2

        ax[i,j].scatter(cal[rvoi]['calib']['cl_suvr'][:,1], cal[rvoi]['calib']['cl_suvr'][:,2], c='black')
        identity_line(ax=ax[i,j], ls='--', c='b')
        m_std = cal[rvoi]['calib']['m_std']
        b_std = cal[rvoi]['calib']['b_std']
        r2 = cal[rvoi]['calib']['r2']
        ax[i,j].text(xystr[rvoi][0], xystr[rvoi][1], f'$y = {m_std:.4f}x + {b_std:.4f}$', fontsize=fontsize)
        ax[i,j].text(xystr[rvoi][0], xystr[rvoi][1]-0.1, f'$R^2={r2:.4f}$', fontsize=fontsize)
        ax[i,j].set_title(rvoi_str[rvoi])
        ax[i,j].set_xlabel('$^\mathrm{PiB}$SUVr$_\mathrm{IND}$', fontsize=fontsize)
        ax[i,j].set_ylabel('$^\mathrm{NEW}$SUVr$_\mathrm{IND}$', fontsize=fontsize)
        ax[i,j].grid('on')

        low_x, high_x = ax[i,j].get_xlim()
        x = np.linspace(low_x, high_x)
        y = m_std*x+b_std
        ax[i,j].plot(x,y,'g')


        ax2[i,j].scatter(cal[rvoi]['calib']['cl_suvr'][:,0], cal[rvoi]['calib']['cl_std_fbb'], c='black')
        identity_line(ax=ax2[i,j], ls='--', c='b')
        ax2[i,j].set_xlabel('CL$^{**}$', fontsize=fontsize)
        ax2[i,j].set_ylabel('$^\mathrm{NEW}$CL$_\mathrm{Std}$', fontsize=fontsize)
        ax2[i,j].grid('on')
        ax2[i,j].set_title(rvoi_str[rvoi])
        # > add the same text about the equation used for converting to PiB SUVrs
        ax[i,j].text(xystr[rvoi][0], xystr[rvoi][1], f'$^{{NEW}}y_{{\mathrm{{SUVr}}}} = {m_std:.4f}^{{PiB}}x_{{\mathrm{{SUVr}}}} + {b_std:.4f}$', fontsize=fontsize)
        ax[i,j].text(xystr[rvoi][0], xystr[rvoi][1]-0.1, f'$R^2={r2:.4f}$', fontsize=fontsize)

        

    fig.tight_layout()
    fig2.tight_layout()
    #----------------------------------------------------------------------

    return cal