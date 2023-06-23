import pickle
import re
import tarfile
import urllib
import xml.etree.ElementTree as ET
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import toml
from niftypet import nimpa
from scipy.stats import linregress

try:          # py<3.9
    import importlib_resources as resources
except ImportError:
    from importlib import resources

try:
    from os import cpu_count
except ImportError:
    try:
        from multiprocessing import cpu_count
    except ImportError:

        def cpu_count():
            return 4


params_path = resources.files('amypet').resolve() / 'params.toml'
params = toml.load(params_path)

# > regions used in CL project
rvois = ['wc', 'cg', 'wcb', 'pns']

# > folder name with the default conversion tables after calibration
cl_anchor_fldr = resources.files('amypet').resolve() / 'data' / 'cl_anchor_tables'
cl_masks_fldr = cl_anchor_fldr.parent / 'CL_masks'
atlas_fldr = cl_anchor_fldr.parent / 'atlas'

# > region full name strings for plots
rvoi_str = {
    'wc': 'WHOLE CEREBELLUM', 'cg': 'CEREBELLUM GM', 'wcb': 'WHOLE CEREBELLUM + BRAIN STEM',
    'pns': 'PONS'}


def is_one_or_more(nargs):
    return nargs and (nargs > 1 if isinstance(nargs, int) else nargs in "+*")


# ----------------------------------------------------------------------
def get_atlas(atlas='aal', res=1):
    '''Get a brain atlas from `neuroparc` out of many available in MNI space.

       Options:
       - atlas:     one of many atlases, e.g., 'aal' (default) or 'hammers';
                    the number of atlases available can be extended by using:
                    'https://github.com/neurodata/neuroparc/raw/master/atlases/label/Human/'
       - res:       the resolution of the atlas in mm.
    '''

    if atlas.lower() == 'aal':
        fatl = atlas_fldr / f'AAL3v1_{res}mm.nii.gz'
        if not fatl.is_file():
            raise IOError('unrecognised atlas!')

        flbl = atlas_fldr / 'AAL3v1.xml'

        tree = ET.parse(flbl)
        lbls = tree.getroot()

        lbls = lbls[1]

        # > atlas dictionary
        datlas = {i[0].text: i[1].text for i in lbls}

        outdct = {'fatlas': fatl, 'flabels': flbl, 'vois': datlas}

    elif atlas.lower() == 'hammers':

        # > main URL
        murl = 'http://biomedic.doc.ic.ac.uk/brain-development/downloads/hammers'

        # > hammers atlas and demographics folders
        hfldr = list(atlas_fldr.glob('Hammers_*n30r95*MNI152*SPM12'))
        dfldr = list(atlas_fldr.glob('Hammers_*ancillary-data'))

        if hfldr and dfldr and (atlas_fldr / 'hammers_license').is_file():
            hfldr = hfldr[0]
            dfldr = dfldr[0]

            fatl = list(hfldr.glob('Hammers_*full*'))
            if len(fatl) != 1:
                raise IOError('Confused with obtaining the Hammers atlas (FULL)')
            else:
                fatl = fatl[0]

            flbl = list(dfldr.glob('Hammers*atlases*n30r95*label*indices*SPM12*.xml'))
            if len(flbl) != 1:
                raise IOError('Confused with obtaining the Hammers atlas labels file')
            else:
                flbl = flbl[0]

        else:

            # > atlas
            urld = urllib.request.urlopen(murl +
                                          '/Hammers_mith-n30r95-maxprob-MNI152-SPM12.tar.gz')
            data = urld.read()

            fatl = atlas_fldr / f'atlases-{atlas}_res-{res}mm.tar.gz'
            with open(fatl, 'wb') as f:
                f.write(data)

            file = tarfile.open(fatl)
            file.extractall(atlas_fldr)
            file.close()

            hfldr = list(atlas_fldr.glob('Hammers_*n30r95*MNI152*SPM12'))
            if len(hfldr) != 1:
                raise IOError('Confused with obtaining the Hammers atlases')
            else:
                hfldr = hfldr[0]

            fatl = list(hfldr.glob('Hammers_*full*'))
            if len(fatl) != 1:
                raise IOError('Confused with obtaining the Hammers atlas (FULL)')
            else:
                fatl = fatl[0]

            # > atlas demographics
            urld = urllib.request.urlopen(murl + '/Hammers_mith-n30-ancillary-data.tar.gz')
            data = urld.read()
            fdmg = atlas_fldr / f'atlas-{atlas}_demographics.tar.gz'
            with open(fdmg, 'wb') as f:
                f.write(data)

            file = tarfile.open(fdmg)
            file.extractall(atlas_fldr)
            file.close()

            dfldr = list(atlas_fldr.glob('Hammers_*ancillary-data'))
            if len(dfldr) != 1:
                raise IOError('Confused with obtaining the Hammers atlas demographics')
            else:
                dfldr = dfldr[0]

            flbl = list(dfldr.glob('Hammers*atlases*n30r95*label*indices*SPM12*.xml'))
            if len(flbl) != 1:
                raise IOError('Confused with obtaining the Hammers atlas labels file')
            else:
                flbl = flbl[0]

        with open(flbl) as f:
            xml = f.read()

        # > correct the first line of xml file if needed
        if xml[:5] != '<?xml':
            xml = '<?xml version="1.0" encoding="UTF-8" ?>\n' + xml

        # > add a single <data> node
        lbls = ET.fromstring(re.sub(r"(<\?xml[^>]+\?>)", r"\1<data>", xml) + "</data>")

        # > atlas dictionary
        datlas = {i[0].text: i[1].text for i in lbls}

        # > dictionary of lobe VOIs
        lobes = ['FL', 'TL', 'PL', 'OL', 'CG', 'in']
        dlobes = {}
        for k in datlas:
            lstr = datlas[k][:2]
            if lstr in lobes:
                if lstr not in dlobes:
                    dlobes[lstr] = [int(k)]
                else:
                    dlobes[lstr].append(int(k))

        if not (atlas_fldr / 'hammers_license').is_file():
            # LICENSE
            import webbrowser
            webbrowser.open('http://brain-development.org/brain-atlases'
                            '/adult-brain-atlases/individual-adult-brain-atlases-new/')
            with open(atlas_fldr / 'hammers_license', 'w') as f:
                f.write('submit the license')

        outdct = {'fatlas': fatl, 'flabels': flbl, 'voi_lobes': dlobes, 'vois': datlas}

    return outdct


def im_check_pairs(fpets, fmris):
    '''
    checks visually image by image if the PET and MR have
    the same orientation.
    '''
    fig, ax = plt.subplots(2, 2, figsize=(10, 10))
    for fp, fm in zip(fpets, fmris):
        print(f'{fp} : {fm}')
        p = nimpa.getnii(fp)
        m = nimpa.getnii(fm)

        ps = p.shape
        ms = m.shape

        ax[0, 0].imshow(p[ps[0] // 2, ...], cmap='magma')
        ax[1, 0].imshow(m[ms[0] // 2, ...], cmap='bone')

        ax[0, 1].imshow(p[..., ps[2] // 2], cmap='magma')
        ax[1, 1].imshow(m[..., ms[2] // 2], cmap='bone')

        plt.draw()
        plt.waitforbuttonpress(0)


# ----------------------------------------------------------------------


# ----------------------------------------------------------------------
def identity_line(ax=None, ls='--', *args, **kwargs):
    '''
    plot identity line for any correlation analysis
    needed for the CL calibration and beyond.
    '''
    # see: https://stackoverflow.com/q/22104256/3986320
    ax = ax or plt.gca()
    identity, = ax.plot([], [], ls=ls, *args, **kwargs)

    def callback(axes):
        low_x, high_x = ax.get_xlim()
        low_y, high_y = ax.get_ylim()
        low = min(low_x, low_y)
        high = max(high_x, high_y)
        identity.set_data([low, high], [low, high])

    callback(ax)
    ax.callbacks.connect('xlim_changed', callback)
    ax.callbacks.connect('ylim_changed', callback)
    return ax


# ----------------------------------------------------------------------


# ----------------------------------------------------------------------
def get_clref(fxls):
    '''
    get the excel sheets with CL reference data from the CL project
    excel sheet.
    '''
    import openpyxl

    # > Define variable to load the workbook
    wb = openpyxl.load_workbook(fxls)
    ws = wb.active

    pib_tbl = {'yc': {}, 'ad': {}}

    # > cell offsets
    ioff_ad = [5, 50]
    ioff_yc = [51, 85]

    sbj = [str(i.value) for i in ws['A']]
    cg = [str(i.value) for i in ws['B']]
    wc = [str(i.value) for i in ws['C']]
    wcb = [str(i.value) for i in ws['D']]
    pns = [str(i.value) for i in ws['E']]

    ccg = [str(i.value) for i in ws['F']]
    cwc = [str(i.value) for i in ws['G']]
    cwcb = [str(i.value) for i in ws['H']]
    cpns = [str(i.value) for i in ws['I']]

    # > young control indices and UR/CL values
    pib_tbl['yc']['id'] = np.array([int(i[3:]) for i in sbj[ioff_yc[0]:ioff_yc[1]]])

    pib_tbl['yc']['ur'] = {}
    pib_tbl['yc']['ur']['cg'] = np.array([float(i) for i in cg[ioff_yc[0]:ioff_yc[1]]])
    pib_tbl['yc']['ur']['wc'] = np.array([float(i) for i in wc[ioff_yc[0]:ioff_yc[1]]])
    pib_tbl['yc']['ur']['wcb'] = np.array([float(i) for i in wcb[ioff_yc[0]:ioff_yc[1]]])
    pib_tbl['yc']['ur']['pns'] = np.array([float(i) for i in pns[ioff_yc[0]:ioff_yc[1]]])

    pib_tbl['yc']['cl'] = {}
    pib_tbl['yc']['cl']['cg'] = np.array([float(i) for i in ccg[ioff_yc[0]:ioff_yc[1]]])
    pib_tbl['yc']['cl']['wc'] = np.array([float(i) for i in cwc[ioff_yc[0]:ioff_yc[1]]])
    pib_tbl['yc']['cl']['wcb'] = np.array([float(i) for i in cwcb[ioff_yc[0]:ioff_yc[1]]])
    pib_tbl['yc']['cl']['pns'] = np.array([float(i) for i in cpns[ioff_yc[0]:ioff_yc[1]]])

    # > AD indices and UR/CL
    pib_tbl['ad']['id'] = np.array([int(i[3:]) for i in sbj[ioff_ad[0]:ioff_ad[1]]])
    pib_tbl['ad']['ur'] = {}
    pib_tbl['ad']['ur']['cg'] = np.array([float(i) for i in cg[ioff_ad[0]:ioff_ad[1]]])
    pib_tbl['ad']['ur']['wc'] = np.array([float(i) for i in wc[ioff_ad[0]:ioff_ad[1]]])
    pib_tbl['ad']['ur']['wcb'] = np.array([float(i) for i in wcb[ioff_ad[0]:ioff_ad[1]]])
    pib_tbl['ad']['ur']['pns'] = np.array([float(i) for i in pns[ioff_ad[0]:ioff_ad[1]]])

    pib_tbl['ad']['cl'] = {}
    pib_tbl['ad']['cl']['cg'] = np.array([float(i) for i in ccg[ioff_ad[0]:ioff_ad[1]]])
    pib_tbl['ad']['cl']['wc'] = np.array([float(i) for i in cwc[ioff_ad[0]:ioff_ad[1]]])
    pib_tbl['ad']['cl']['wcb'] = np.array([float(i) for i in cwcb[ioff_ad[0]:ioff_ad[1]]])
    pib_tbl['ad']['cl']['pns'] = np.array([float(i) for i in cpns[ioff_ad[0]:ioff_ad[1]]])

    return pib_tbl
    # ----------------------------------------------------------------------


def check_urs(ur_yc, ur_ad, refs):
    ''' Check/QC the obtained URs in `ur_yc` and `ur_ad `dictionaries
        in relation to the reference in `refs` dictionary for the groups
        ('ad' or 'yc').

        Returns dictionary with the mean URs and differences.
    '''

    # > prepare diff dictionary for the differences observed
    # > in young controls (yc) and AD patients (ad)
    diff = {'yc': {}, 'ad': {}}

    # > initialise the means for each reference VOI
    for rvoi in rvois:
        diff['yc'][rvoi] = {'mean_ref': 0, 'mean': 0, 'N': 0}
        diff['ad'][rvoi] = {'mean_ref': 0, 'mean': 0, 'N': 0}

    def run_checks(ur_dct, grp):

        for rvoi in rvois:
            print('========================================================')
            for k in ur_dct:

                if grp == 'yc':
                    idx = int(k[2:5])
                elif grp == 'ad':
                    idx = int(k[2:4])
                else:
                    raise ValueError('e> unknown group - only <yc>  or <ad> are accepted')

                i = np.where(refs[grp]['id'] == idx)[0][0]

                ur = ur_dct[k]['ur'][rvoi]
                ur_ref = refs[grp]['ur'][rvoi][i]
                err = 100 * (ur-ur_ref) / ur_ref

                diff[grp][rvoi]['N'] += 1
                diff[grp][rvoi]['mean_ref'] += ur_ref
                diff[grp][rvoi]['mean'] += ur
                diff[grp][rvoi][k] = {'ur': ur, 'ref': ur_ref, 'err': err}

                print(f'refvoi={rvoi}, indx> {idx}, ur={ur:.3f},'
                      f' ref={ur_ref:.3f}, error={err:.3f}%')

            diff[grp][rvoi]['mean'] /= diff[grp][rvoi]['N']
            diff[grp][rvoi]['mean_ref'] /= diff[grp][rvoi]['N']

            # relative % mean difference
            emean = diff[grp][rvoi]['mean']
            rmean = diff[grp][rvoi]['mean_ref']
            rmd = (emean-rmean) / rmean
            diff[grp][rvoi]['mean_diff'] = rmd
            print('--------------------------------------------------------')
            print(f'> group % mean difference: {rmd:.3f}% (ur={emean:.3f}, ref={rmean:.3f})')

    print('========================================================')
    print('YOUNG CONTROLS (YC)')
    print('========================================================')
    run_checks(ur_yc, 'yc')

    print('========================================================')
    print('AD PATIENTS (AD)')
    print('========================================================')
    run_checks(ur_ad, 'ad')

    return diff


# ----------------------------------------------------------------------


# ----------------------------------------------------------------------
def check_cls(ur_yc, ur_ad, diff, refs):
    ''' check centiloid values relative to the reference in `refs`
        dictionary, using UR dictionaries for YC and AD groups
        `ur_yc` and `ur_ad`.  Also, update the `diff` dictionary
        with the centiloid results and output it.
    '''

    for rvoi in rvois:
        print('=====================================================')
        ur_0 = diff['yc'][rvoi]['mean']
        ur_100 = diff['ad'][rvoi]['mean']

        grp = 'yc'
        ur_dct = ur_yc

        for k in ur_dct:

            idx = int(k[2:5])
            i = np.where(refs[grp]['id'] == idx)[0][0]

            cl = 100 * (ur_dct[k]['ur'][rvoi] - ur_0) / (ur_100-ur_0)
            cl_ref = refs[grp]['cl'][rvoi][i]

            diff[grp][rvoi][k].update(cl=cl, cl_ref=cl_ref)

            print(f'refvoi={rvoi}, indx> {idx}, cl={cl:.3f}, ref={cl_ref:.3f}')

        grp = 'ad'
        ur_dct = ur_ad

        for k in ur_dct:

            idx = int(k[2:4])
            i = np.where(refs[grp]['id'] == idx)[0][0]

            cl = 100 * (ur_dct[k]['ur'][rvoi] - ur_0) / (ur_100-ur_0)
            cl_ref = refs[grp]['cl'][rvoi][i]

            diff[grp][rvoi][k].update(cl=cl, cl_ref=cl_ref)

            print(f'refvoi={rvoi}, indx> {idx}, cl={cl:.3f}, ref={cl_ref:.3f}')

    def get_cls(diff, grp, rvoi, cl='cl'):
        ''' Get centiloid values in a list for any group and reference VOI.
        '''
        return [diff[grp][rvoi][k][cl] for k in diff[grp][rvoi] if grp in k.lower()]

    # ------------------------------------------------------------------
    fig, ax = plt.subplots(2, 2, figsize=(10, 10))

    for a, rvoi in enumerate(rvois):
        i = a // 2
        j = a % 2

        # > combine CLs for both groups
        cl = get_cls(diff, 'yc', rvoi, cl='cl')
        cl += get_cls(diff, 'ad', rvoi, cl='cl')

        # > combine reference CLs into one list
        clref = get_cls(diff, 'yc', rvoi, cl='cl_ref')
        clref += get_cls(diff, 'ad', rvoi, cl='cl_ref')

        # > find the linear regression between the CL and reference CL
        m, a, r, p, stderr = linregress(clref, cl)
        r2 = r**2

        # > plot the scatter plots for all 4 reference VOIs
        ax[i, j].scatter(clref, cl, c='black')
        identity_line(ax=ax[i, j], ls='--', c='b')
        ax[i, j].text(60, 20, f'$y = {m:.4f}x + {a:.4f}$', fontsize=12)
        ax[i, j].text(60, 10, f'$R^2={r2:.4f}$', fontsize=12)
        ax[i, j].set_title(rvoi_str[rvoi])
        ax[i, j].set_xlabel('CL ref')
        ax[i, j].set_ylabel('CL AmyPET')
        ax[i, j].grid('on')

    fig.tight_layout()
    # ------------------------------------------------------------------

    return diff


# ----------------------------------------------------------------------

# ======================================================================
# CL CALIBRATION FOR A NEW TRACER
# ======================================================================


def calib_tracer(outpib, outnew, xystr=None, figsize=(10, 10), fontsize=12):
    '''calibrate CL scale for a new amyloid PET tracer,
       given the dictionaries from processing the PiB
       and a new tracer scans, i.e., `outpib` and `outnew`.

       Other arguments:
       xystr - dictionary of plot string XY locations.
       figsize - tuple for controlling the figure size
    '''

    # > get PiB centiloid anchor points
    clap = get_cl_anchors()

    if xystr is None:
        xystr = {'wc': [1.5, 1.1], 'cg': [2.0, 1.2], 'wcb': [1.5, 1.0], 'pns': [1.0, 0.75]}

    # > extracting the index from names
    pp = re.compile(r'(GE_\w*_\w*_(\d*)_NIFTI)|(Y*(\d*)\w*)')

    # > calibration dictionary with UR and CL values
    cal = {}
    for rvoi in rvois:

        print('------------------------------------------------------------')
        print(rvoi_str[rvoi])

        # > PiB and F18 URs and PiB CL in one list/array
        cl_ur = []

        # > initialise dictionary
        cal[rvoi] = {'sbj': {}, 'calib': {}}

        for k in outpib:

            idx = pp.match(k)[2] or pp.match(k)[4]

            # > get the key name in the NEW tracer dataset
            # > (matching using the first 20 characters)
            kf = [ky for ky in outnew if idx in ky[:20]] # idx+'_'
            if len(kf) == 1:
                kf = kf[0]
            else:
                raise ValueError('e> problem with F18 index...')

            # > get the UR for PiB and FBB
            urp = outpib[k]['ur'][rvoi]
            urf = outnew[kf]['ur'][rvoi]

            # > calculate the CL for the calibrating PiB scans
            cl = 100 * (urp - clap[rvoi][0]) / (clap[rvoi][1] - clap[rvoi][0])

            # > put the URs for PiB and FBB into dictionary
            cal[rvoi]['sbj'][idx] = {'cl': cl, 'urp': urp, 'urf': urf}

            # > ... and a list for a future array
            cl_ur.append([cl, cal[rvoi]['sbj'][idx]['urp'], cal[rvoi]['sbj'][idx]['urf']])

            print(f'refvoi={rvoi}, indx> {idx}, cl={cl:.3f},'
                  f' ur_pib={urp:.3f}, ur_fbb={urf:.3f}')

        # --------------------------------------------------------------
        # > find the linear relationship for FBB_UR = m*PiB_UR + b
        # > calculate the tracer `m_std` and `b_std`
        # > (Eq. 2.2.3.1a in Klunk et al. 2015)

        cl_ur = np.array(cl_ur)
        m_std, b_std, r, p, stderr = linregress(cl_ur[:, 1], cl_ur[:, 2])
        r2 = r**2

        ur_pib_calc = (cl_ur[:, 2] - b_std) / m_std

        cl_std_fbb = 100 * (ur_pib_calc - clap[rvoi][0]) / (clap[rvoi][1] - clap[rvoi][0])

        cal[rvoi]['calib'] = {
            'm_std': m_std, 'b_std': b_std, 'r2': r2, 'p': p, 'stderr': stderr, 'cl_ur': cl_ur,
            'ur_pib_calc': ur_pib_calc, 'cl_std_fbb': cl_std_fbb}
        # --------------------------------------------------------------

    # ------------------------------------------------------------------
    # VISUALISATION

    fig, ax = plt.subplots(2, 2, figsize=figsize)
    fig2, ax2 = plt.subplots(2, 2, figsize=figsize)

    for a, rvoi in enumerate(rvois):
        i = a // 2
        j = a % 2

        ax[i, j].scatter(cal[rvoi]['calib']['cl_ur'][:, 1], cal[rvoi]['calib']['cl_ur'][:, 2],
                         c='black')
        identity_line(ax=ax[i, j], ls='--', c='b')
        m_std = cal[rvoi]['calib']['m_std']
        b_std = cal[rvoi]['calib']['b_std']
        r2 = cal[rvoi]['calib']['r2']
        ax[i, j].text(xystr[rvoi][0], xystr[rvoi][1], f'$y = {m_std:.4f}x + {b_std:.4f}$',
                      fontsize=fontsize)
        ax[i, j].text(xystr[rvoi][0], xystr[rvoi][1] - 0.1, f'$R^2={r2:.4f}$', fontsize=fontsize)
        ax[i, j].set_title(rvoi_str[rvoi], fontweight='bold')
        ax[i, j].set_xlabel(r'$^\mathrm{PiB}$UR$_\mathrm{IND}$', fontsize=fontsize)
        ax[i, j].set_ylabel(r'$^\mathrm{NEW}$UR$_\mathrm{IND}$', fontsize=fontsize)
        ax[i, j].grid('on')

        low_x, high_x = ax[i, j].get_xlim()
        x = np.linspace(low_x, high_x)
        y = m_std*x + b_std
        ax[i, j].plot(x, y, 'g')

        ax2[i, j].scatter(cal[rvoi]['calib']['cl_ur'][:, 0], cal[rvoi]['calib']['cl_std_fbb'],
                          c='black')
        identity_line(ax=ax2[i, j], ls='--', c='b')
        ax2[i, j].set_xlabel('CL$^{**}$', fontsize=fontsize)
        ax2[i, j].set_ylabel(r'$^\mathrm{NEW}$CL$_\mathrm{Std}$', fontsize=fontsize)
        ax2[i, j].grid('on')
        ax2[i, j].set_title(rvoi_str[rvoi])
        # > add the same text about the equation used for converting to PiB URs
        ax2[i, j].text(
            40, 10, f'$y_{{\\mathrm{{UR}}}}^{{\\mathrm{{NEW}}}}'
            f' = {m_std:.4f}x_{{\\mathrm{{UR}}}}^{{\\mathrm{{PiB}}}} + {b_std:.4f}$',
            fontsize=fontsize)
        ax2[i, j].text(40, -5, f'$R^2={r2:.4f}$', fontsize=fontsize)

    fig.tight_layout()
    fig2.tight_layout()
    # ------------------------------------------------------------------

    return cal


# ----------------------------------------------------------------------
def save_cl_anchors(diff, outpath=None):
    """save the CL anchor points for each reference VOI."""

    CLA = {}
    txt = '# centiloid anchor points for different reference VOIs\n'
    for rvoi in rvois:
        ur_0 = diff['yc'][rvoi]['mean']
        ur_100 = diff['ad'][rvoi]['mean']
        CLA[rvoi] = (ur_0, ur_100)
        txt += f'cla_{rvoi} = ({ur_0}, {ur_100})\n'
    print(txt)

    if outpath is None:
        outpath = cl_anchor_fldr

    outpath = Path(outpath) / 'CL_PiB_anchors.pkl'
    nimpa.create_dir(outpath.parent)
    with open(outpath, 'wb') as f:
        pickle.dump(CLA, f)

    CLA['path'] = outpath
    return CLA


def get_cl_anchors(path=None):
    """load the centiloid anchor points"""
    if path is None:
        path = cl_anchor_fldr
    path = Path(path) / 'CL_PiB_anchors.pkl'
    with open(path, 'rb') as f:
        return pickle.load(f)


# ----------------------------------------------------------------------


# ----------------------------------------------------------------------
def save_ur2pib(cal, tracer, outpath=None):
    ''' save the linear transformation parameters, `m_Std` and `b_Std`;
        the parameters are taken from the bigger calibration dictionary `cal`
        used here as input.

        The compact dictionary consists the two parameters for all reference regions
        and for the specific F-18 tracer being used.

        `tracer` can be:
            - 'fbb' for [18F]florbetaben
            - 'flute' for [18F]flutemetamol
            - 'fbp' for [18F]florbetapir
    '''

    if tracer not in ('fbp', 'fbb', 'flute'):
        raise ValueError('tracer is unrecognised or not given!')

    # conversion dictionary
    CNV = {}
    txt = '# Tracer uptake ratio (UR) conversion to PiB UR\n'
    for rvoi in rvois:
        mstd = cal[rvoi]['calib']['m_std']
        bstd = cal[rvoi]['calib']['b_std']
        CNV[rvoi] = {'m_std': mstd, 'b_std': bstd}
        txt += f'{rvoi}: m_std, b_std = ({mstd}, {bstd})\n'
    print(txt)

    if outpath is None:
        outpath = cl_anchor_fldr
    outpath = Path(outpath)

    nimpa.create_dir(outpath)
    pth = outpath / f'ur_{tracer}_to_ur_pib__transform.pkl'
    with open(pth, 'wb') as f:
        pickle.dump(CNV, f)

    CNV['path'] = pth
    return CNV


def get_ur2pib(tracer, path=None):
    """
    load the linear transformation parameters from a tracer uptake
    ratio (UR) to PiB UR; it's used for converting to CL scale any F-18
    tracers.
    """
    if tracer not in ['fbp', 'fbb', 'flute']:
        raise ValueError(f'tracer ({tracer}) unrecognised')

    if path is None:
        path = cl_anchor_fldr
    pth = Path(path) / f'ur_{tracer}_to_ur_pib__transform.pkl'

    with open(pth, 'rb') as f:
        return pickle.load(f)
